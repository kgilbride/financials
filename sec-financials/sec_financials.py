#!/usr/bin/env python3
"""
SEC Financial Statement Puller — engine (edgartools)
====================================================
Pulls a company's AS-REPORTED income statement, balance sheet, and cash flow
from its latest 10-K (SEC EDGAR XBRL) and writes an Excel workbook.

Two views per statement:
  • "<Statement>"          clean model view: face lines, $ millions, latest 3 FYs
  • "<Statement> (detail)" full as-reported detail, all periods, full dollars

Use from the command line:
    python sec_financials.py GTM
…or import and call generate("GTM", out_dir=..., status=print) from the app.
"""

import os
import re
import sys
import pandas as pd

IDENTITY = os.environ.get("EDGAR_IDENTITY", "")  # the user supplies their own name + email
META_COLS = ("concept", "label", "standard_concept", "level", "abstract", "dimension", "units", "decimals")


# ---------- data extraction ----------
def _statement_df(fin, method_names):
    for name in method_names:
        fn = getattr(fin, name, None)
        if fn is None:
            continue
        try:
            stmt = fn()
        except Exception:
            continue
        if stmt is None:
            continue
        for conv in ("to_dataframe", "to_pandas", "get_dataframe"):
            c = getattr(stmt, conv, None)
            if callable(c):
                try:
                    df = c()
                    if df is not None and len(df):
                        return df
                except Exception:
                    pass
        try:
            df = pd.DataFrame(stmt)
            if len(df):
                return df
        except Exception:
            pass
    return None


def _period_columns(df):
    # A period column's header contains a date (YYYY-MM-DD) or an (FY) tag.
    return [c for c in df.columns if re.search(r"\d{4}-\d{2}-\d{2}", str(c)) or "(FY)" in str(c)]


def _year_key(c):
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", str(c))
    return m.group(0) if m else str(c)


def _latest_annual_columns(df, n=3):
    pcols = _period_columns(df)
    fy = [c for c in pcols if "(FY)" in str(c)]
    cols = fy if fy else pcols
    return sorted(cols, key=_year_key)[-n:]


def _fy_label(c):
    m = re.search(r"(\d{4})-\d{2}-\d{2}", str(c))
    return "FY'" + m.group(1)[2:] + "a" if m else str(c)


def _scale(concept, v):
    if v is None or (isinstance(v, float) and pd.isna(v)) or v == "":
        return None
    if not isinstance(v, (int, float)):
        return v
    return round(v, 2) if "PerShare" in str(concept) else v / 1e6


def _build_clean(df):
    cols = _latest_annual_columns(df, 3)
    seen, keep = set(), []
    for i, row in df.iterrows():
        con = row.get("concept")
        if con in seen:
            continue
        seen.add(con)
        keep.append(i)
    sub = df.loc[keep]
    out = pd.DataFrame()
    out["Line item"] = sub["label"].astype(str)
    out["_concept"] = sub["concept"].astype(str)
    for c in cols:
        out[_fy_label(c)] = [_scale(con, v) for con, v in zip(sub["concept"], sub[c])]
    return out


def _build_detail(df):
    keep = [c for c in ("concept", "label") if c in df.columns] + _period_columns(df)
    return df[keep].copy()


# ---------- workbook writing ----------
_BOLD = {"Gross profit", "Total operating expenses", "Income from operations", "Total current assets",
         "Total assets", "Total current liabilities", "Total liabilities", "Total stockholders' equity",
         "Total liabilities and stockholders' equity", "Net income"}


def _write_clean(wb, name, out):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws = wb.create_sheet(name[:31])
    concepts = out["_concept"].tolist()
    view = out.drop(columns=["_concept"])
    ws.append(list(view.columns))
    for _, row in view.iterrows():
        ws.append([None if pd.isna(v) else v for v in row.tolist()])
    ncol = ws.max_column
    navy = PatternFill("solid", fgColor="1F3864")
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = navy
        cell.alignment = Alignment(horizontal="left" if c == 1 else "center")
    for r in range(2, ws.max_row + 1):
        con = concepts[r - 2]
        eps = "PerShare" in con
        label = ws.cell(row=r, column=1).value
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Arial", bold=(label in _BOLD))
            if c > 1 and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00" if eps else '#,##0.0;(#,##0.0);"-"'
    ws.column_dimensions["A"].width = 52
    for c in range(2, ncol + 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 12
    ws.freeze_panes = "B2"


def _write_detail(wb, name, df):
    from openpyxl.styles import Font, PatternFill
    ws = wb.create_sheet(name[:31])
    grey = PatternFill("solid", fgColor="595959")
    ws.append([str(c) for c in df.columns])
    for _, row in df.iterrows():
        ws.append([None if pd.isna(v) else v for v in row.tolist()])
    ncol = ws.max_column
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = grey
    for r in range(2, ws.max_row + 1):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Arial")
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0;(#,##0);"-"'
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 50
    for c in range(3, ncol + 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 16
    ws.freeze_panes = "C2"


def _source_info(company):
    try:
        f = company.get_filings(form="10-K").latest()
        return f"Form 10-K, accession {f.accession_no}, filed {f.filing_date}"
    except Exception:
        return "Latest 10-K (SEC EDGAR XBRL)"


# ---------- public entry point ----------
def generate(ticker, out_dir=None, identity=IDENTITY, status=lambda m: None):
    """Pull `ticker` and write <TICKER>_financials.xlsx into out_dir.
    Returns the path to the .xlsx. `status` is a callback for progress messages."""
    ticker = ticker.strip().upper()
    out_dir = out_dir or os.getcwd()
    os.makedirs(out_dir, exist_ok=True)

    ident = (identity or IDENTITY or "").strip()
    if "@" not in ident:
        raise RuntimeError("Please enter your name and email (SEC requires it), "
                           "e.g. 'Jane Smith jane@company.com'.")

    from edgar import Company, set_identity
    from openpyxl import Workbook
    from openpyxl.styles import Font

    set_identity(ident)
    status(f"Resolving {ticker} on SEC EDGAR…")
    company = Company(ticker)
    name = getattr(company, "name", ticker)
    cik = getattr(company, "cik", "")
    status(f"Found {name} (CIK {cik}). Downloading latest 10-K financials…")

    fin = company.get_financials()
    if fin is None:
        raise RuntimeError(f"No XBRL financials found for {ticker}.")

    raw = {
        "Income Statement": _statement_df(fin, ["income_statement", "income", "get_income_statement"]),
        "Balance Sheet":    _statement_df(fin, ["balance_sheet", "get_balance_sheet"]),
        "Cash Flow":        _statement_df(fin, ["cashflow_statement", "cash_flow_statement", "cashflow", "cash_flow", "get_cash_flow_statement"]),
    }

    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    for i, (text, size, color, bold) in enumerate([
        (f"{name} ({ticker})", 14, "1F3864", True),
        ("Financial statements — as reported from SEC EDGAR XBRL", 10, "595959", False),
        ("", 10, "000000", False),
        (f"Entity: {name}", 10, "000000", False),
        (f"SEC CIK: {cik}", 10, "000000", False),
        (f"Source: {_source_info(company)}", 10, "000000", False),
        ("Clean tabs: face lines, $ millions, latest 3 FYs.  Detail tabs: all lines, full $.", 10, "595959", False),
    ], start=1):
        cover.cell(row=i, column=1, value=text).font = Font(name="Arial", size=size, bold=bold, color=color)
    cover.column_dimensions["A"].width = 82

    for sheet, df in raw.items():
        if df is None or not len(df):
            wb.create_sheet(sheet)["A1"] = "(not available for this filer)"
            status(f"  {sheet}: not available")
            continue
        clean = _build_clean(df)
        detail = _build_detail(df)
        _write_clean(wb, sheet, clean)
        _write_detail(wb, f"{sheet} (detail)", detail)
        status(f"  {sheet}: {len(clean)} clean lines, {len(detail)} detail lines")

    path = os.path.join(out_dir, f"{ticker}_financials.xlsx")
    wb.save(path)
    status(f"Saved {os.path.basename(path)}")
    return path


def main():
    ticker = (sys.argv[1] if len(sys.argv) > 1 else input("Ticker: ")).strip().upper()
    if not ticker:
        print("No ticker given."); return
    path = generate(ticker, out_dir=os.getcwd(), status=print)
    print(f"\nDone: {path}")
    print("Open in Google Sheets with:  File > Import > Upload")


if __name__ == "__main__":
    main()
